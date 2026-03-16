"""
Metadata Cache for Data Analysis V3
Efficiently handles metadata extraction and caching for large files
"""

import os
import json
import logging
import re
import pandas as pd
from .encoding_handler import EncodingHandler
from typing import Dict, Any, Optional, Tuple
from datetime import datetime
from pandas.api.types import is_numeric_dtype

logger = logging.getLogger(__name__)


class MetadataCache:
    """
    Handles metadata caching for uploaded files.
    Prevents loading entire large files just to get basic info.
    """
    
    # Sampling thresholds
    SMALL_FILE_SIZE = 10 * 1024 * 1024  # 10MB
    SAMPLE_ROWS = 1000  # Sample size for large files
    
    @staticmethod
    def get_cache_path(session_id: str) -> str:
        """Get the path to the metadata cache file."""
        return f"instance/uploads/{session_id}/metadata_cache.json"
    
    @staticmethod
    def load_cache(session_id: str) -> Optional[Dict[str, Any]]:
        """Load cached metadata if it exists."""
        cache_path = MetadataCache.get_cache_path(session_id)
        if os.path.exists(cache_path):
            try:
                with open(cache_path, 'r') as f:
                    return json.load(f)
            except Exception as e:
                logger.error(f"Error loading metadata cache: {e}")
        return None
    
    @staticmethod
    def save_cache(session_id: str, metadata: Dict[str, Any]):
        """Save metadata to cache."""
        cache_path = MetadataCache.get_cache_path(session_id)
        try:
            os.makedirs(os.path.dirname(cache_path), exist_ok=True)
            with open(cache_path, 'w') as f:
                json.dump(metadata, f, indent=2, default=str)
            logger.info(f"Metadata cache saved for session {session_id}")
        except Exception as e:
            logger.error(f"Error saving metadata cache: {e}")
    
    @staticmethod
    def extract_file_metadata(filepath: str, filename: str) -> Dict[str, Any]:
        """
        Extract metadata from a file without loading it entirely if large.
        
        Args:
            filepath: Full path to the file
            filename: Original filename
            
        Returns:
            Dictionary containing file metadata
        """
        metadata = {
            'filename': filename,
            'filepath': filepath,
            'file_size': os.path.getsize(filepath),
            'file_size_mb': round(os.path.getsize(filepath) / (1024 * 1024), 2),
            'upload_time': datetime.now().isoformat(),
            'is_sampled': False
        }
        
        try:
            if filename.endswith('.csv'):
                metadata.update(MetadataCache._extract_csv_metadata(filepath))
            elif filename.endswith(('.xlsx', '.xls')):
                metadata.update(MetadataCache._extract_excel_metadata(filepath))
            elif filename.endswith('.json'):
                metadata.update(MetadataCache._extract_json_metadata(filepath))
            else:
                metadata['type'] = 'text'
                metadata['rows'] = 'N/A'
                metadata['columns'] = 'N/A'
        except Exception as e:
            logger.error(f"Error extracting metadata from {filename}: {e}")
            metadata['error'] = str(e)
            metadata['rows'] = 'Error'
            metadata['columns'] = 'Error'
        
        return metadata
    
    @staticmethod
    def _extract_csv_metadata(filepath: str) -> Dict[str, Any]:
        """Extract metadata from CSV file."""
        file_size = os.path.getsize(filepath)
        
        if file_size <= MetadataCache.SMALL_FILE_SIZE:
            # Small file: load fully
            # Check if this is TPR data to decide on sanitization
            df_test = pd.read_csv(filepath, nrows=5)
            is_tpr_data = any('RDT' in col or 'Microscopy' in col or 'TPR' in col for col in df_test.columns)
            
            df = EncodingHandler.read_csv_with_encoding(filepath)
            return {
                'type': 'csv',
                'rows': df.shape[0],
                'columns': df.shape[1],
                'column_names': list(df.columns),
                'dtypes': {col: str(dtype) for col, dtype in df.dtypes.items()},
                'is_sampled': False
            }
        else:
            # Large file: sample first rows
            # Check if this is TPR data to decide on sanitization
            df_test = pd.read_csv(filepath, nrows=5)
            is_tpr_data = any('RDT' in col or 'Microscopy' in col or 'TPR' in col for col in df_test.columns)
            
            df_sample = EncodingHandler.read_csv_with_encoding(filepath, nrows=MetadataCache.SAMPLE_ROWS)
            
            # Estimate total rows from file size
            sample_size = df_sample.memory_usage(deep=True).sum()
            estimated_rows = int((file_size / sample_size) * MetadataCache.SAMPLE_ROWS)
            
            return {
                'type': 'csv',
                'rows': estimated_rows,
                'rows_estimated': True,
                'columns': df_sample.shape[1],
                'column_names': list(df_sample.columns),
                'dtypes': {col: str(dtype) for col, dtype in df_sample.dtypes.items()},
                'is_sampled': True,
                'sample_rows': MetadataCache.SAMPLE_ROWS
            }
    
    @staticmethod
    def _extract_excel_metadata(filepath: str) -> Dict[str, Any]:
        """Extract metadata from Excel file."""
        file_size = os.path.getsize(filepath)
        
        if file_size <= MetadataCache.SMALL_FILE_SIZE:
            # Small file: load fully
            df = pd.read_excel(filepath)
            return {
                'type': 'excel',
                'rows': df.shape[0],
                'columns': df.shape[1],
                'column_names': list(df.columns),
                'dtypes': {col: str(dtype) for col, dtype in df.dtypes.items()},
                'is_sampled': False
            }
        else:
            # Large file: sample first rows
            # For Excel, we can use nrows parameter
            df_sample = pd.read_excel(filepath, nrows=MetadataCache.SAMPLE_ROWS)
            
            # For Excel files, we can try to get exact row count without loading all data
            # Using openpyxl to get dimensions
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filepath, read_only=True, data_only=True)
                ws = wb.active
                total_rows = ws.max_row
                wb.close()
                
                return {
                    'type': 'excel',
                    'rows': total_rows,
                    'rows_exact': True,
                    'columns': df_sample.shape[1],
                    'column_names': list(df_sample.columns),
                    'dtypes': {col: str(dtype) for col, dtype in df_sample.dtypes.items()},
                    'is_sampled': True,
                    'sample_rows': MetadataCache.SAMPLE_ROWS
                }
            except:
                # Fallback to estimation
                sample_size = df_sample.memory_usage(deep=True).sum()
                # Excel files have overhead, adjust estimation
                estimated_rows = int((file_size / sample_size) * MetadataCache.SAMPLE_ROWS * 0.7)
                
                return {
                    'type': 'excel',
                    'rows': estimated_rows,
                    'rows_estimated': True,
                    'columns': df_sample.shape[1],
                    'column_names': list(df_sample.columns),
                    'dtypes': {col: str(dtype) for col, dtype in df_sample.dtypes.items()},
                    'is_sampled': True,
                    'sample_rows': MetadataCache.SAMPLE_ROWS
                }
    
    @staticmethod
    def _extract_json_metadata(filepath: str) -> Dict[str, Any]:
        """Extract metadata from JSON file."""
        with open(filepath, 'r') as f:
            data = json.load(f)
        
        if isinstance(data, list):
            return {
                'type': 'json',
                'rows': len(data),
                'structure': 'array',
                'is_sampled': False
            }
        elif isinstance(data, dict):
            return {
                'type': 'json',
                'keys': list(data.keys()),
                'structure': 'object',
                'is_sampled': False
            }
        else:
            return {
                'type': 'json',
                'structure': type(data).__name__,
                'is_sampled': False
            }
    
    @staticmethod
    def update_file_metadata(session_id: str, filepath: str, filename: str):
        """
        Update metadata cache for a specific file.
        
        Args:
            session_id: User session ID
            filepath: Full path to the file
            filename: Original filename
        """
        # Load existing cache or create new
        cache = MetadataCache.load_cache(session_id) or {'files': {}}
        
        # Extract metadata for this file
        metadata = MetadataCache.extract_file_metadata(filepath, filename)
        
        # Build dataset profile (pre-computed overview)
        profile = MetadataCache.build_dataset_profile(filepath, metadata)
        if profile:
            metadata['profile'] = profile

        # Update cache
        cache['files'][filename] = metadata
        cache['last_updated'] = datetime.now().isoformat()

        # Save cache
        MetadataCache.save_cache(session_id, cache)
        
        return metadata
    
    @staticmethod
    def get_file_metadata(session_id: str, filename: Optional[str] = None) -> Optional[Dict[str, Any]]:
        """Retrieve cached metadata for the latest uploaded file or a specific filename."""
        cache = MetadataCache.load_cache(session_id)
        if not cache:
            return None
        files = cache.get('files') or {}
        if not files:
            return None
        if filename:
            return files.get(filename)
        # Return the most recently cached file (dict preserves insertion order in Python 3.7+)
        try:
            last_filename = next(reversed(files))
            return files.get(last_filename)
        except StopIteration:
            return None
    
    @staticmethod
    def get_summary_from_cache(session_id: str) -> Optional[str]:
        """
        Get a formatted summary of available data from cache.
        
        Args:
            session_id: User session ID
            
        Returns:
            Formatted string summary or None if no cache
        """
        cache = MetadataCache.load_cache(session_id)
        if not cache or 'files' not in cache:
            return None
        
        summary_parts = ["Available data:"]
        
        for filename, metadata in cache['files'].items():
            if metadata.get('error'):
                summary_parts.append(f"- {filename}: Error loading file")
            elif metadata.get('type') in ['csv', 'excel'] and 'rows' in metadata:
                rows = metadata.get('rows', 'Unknown')
                cols = metadata.get('columns', 'Unknown')
                
                # Format row count
                if isinstance(rows, int):
                    if metadata.get('rows_estimated'):
                        rows_str = f"~{rows:,}"
                    else:
                        rows_str = f"{rows:,}"
                else:
                    rows_str = str(rows)
                
                # Add size info for large files
                size_mb = metadata.get('file_size_mb', 0)
                if size_mb > 10:
                    size_info = f" ({size_mb:.1f}MB)"
                else:
                    size_info = ""
                
                var_name = filename.split('.')[0].replace(' ', '_').replace('-', '_')
                summary_parts.append(f"- {var_name}: {rows_str} rows, {cols} columns{size_info}")
                
                # Add sampling note if applicable
                if metadata.get('is_sampled'):
                    summary_parts.append(f"  (Metadata from first {metadata.get('sample_rows', 1000)} rows)")
            elif metadata.get('type') == 'json':
                structure = metadata.get('structure', 'Unknown')
                if structure == 'array':
                    summary_parts.append(f"- {filename}: JSON array with {metadata.get('rows', 'Unknown')} items")
                else:
                    summary_parts.append(f"- {filename}: JSON {structure}")
            else:
                summary_parts.append(f"- {filename}: {metadata.get('type', 'Unknown')} file")
        
        return "\n".join(summary_parts)

    @staticmethod
    def get_file_profile(session_id: str, filename: Optional[str]) -> Optional[Dict[str, Any]]:
        if not filename:
            return None
        cache = MetadataCache.load_cache(session_id)
        if not cache:
            return None
        return cache.get('files', {}).get(filename, {}).get('profile')

    @staticmethod
    def build_dataset_profile(filepath: str, metadata: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Generate a lightweight dataset profile for quick responses."""
        file_type = metadata.get('type')
        if file_type not in {'csv', 'excel'}:
            return None

        try:
            if file_type == 'csv':
                sample = EncodingHandler.read_csv_with_encoding(filepath, nrows=min(MetadataCache.SAMPLE_ROWS, 5000))
            else:
                sample = pd.read_excel(filepath, nrows=min(MetadataCache.SAMPLE_ROWS, 5000))
        except Exception as e:
            logger.warning(f"Dataset profile: unable to read sample from {filepath}: {e}")
            return None

        if sample is None or sample.empty:
            return None

        metrics = MetadataCache._summarize_dataframe(sample, metadata)
        overview_lines = MetadataCache._build_overview_lines(metadata, metrics)

        return {
            'overview_text': "\n".join(overview_lines),
            'metrics': metrics,
            'generated_at': datetime.now().isoformat()
        }

    @staticmethod
    def _summarize_dataframe(df: pd.DataFrame, metadata: Dict[str, Any]) -> Dict[str, Any]:
        metrics: Dict[str, Any] = {}

        rows = metadata.get('rows')
        metrics['rows'] = rows
        metrics['rows_estimated'] = metadata.get('rows_estimated', False)
        metrics['columns'] = metadata.get('columns', df.shape[1])
        metrics['column_examples'] = list(df.columns[:5])  # Show first 5 only

        numeric_cols = [col for col in df.columns if is_numeric_dtype(df[col])]
        categorical_cols = [col for col in df.columns if df[col].dtype == object and df[col].nunique(dropna=True) < 50]

        metrics['numeric_columns'] = numeric_cols[:5]
        metrics['categorical_columns'] = categorical_cols[:5]

        # Value distributions for categorical columns (exact values the LLM needs)
        value_distributions: Dict[str, Any] = {}
        for col in categorical_cols:
            unique_vals = df[col].dropna().unique()
            n_unique = len(unique_vals)
            if n_unique <= 25:
                value_distributions[col] = {
                    'unique_count': n_unique,
                    'values': sorted(str(v) for v in unique_vals),
                }
            else:
                top5 = df[col].value_counts(dropna=True).head(5)
                value_distributions[col] = {
                    'unique_count': n_unique,
                    'top_values': {str(k): int(v) for k, v in top5.items()},
                }
        metrics['value_distributions'] = value_distributions

        # Numeric ranges
        numeric_ranges: Dict[str, Dict[str, float]] = {}
        for col in numeric_cols:
            col_data = df[col].dropna()
            if len(col_data) > 0:
                numeric_ranges[col] = {
                    'min': float(col_data.min()),
                    'max': float(col_data.max()),
                    'mean': round(float(col_data.mean()), 2),
                }
        metrics['numeric_ranges'] = numeric_ranges

        # Data type summary
        dtype_counts = df.dtypes.value_counts()
        metrics['dtype_summary'] = {str(dtype): int(count) for dtype, count in dtype_counts.items()}

        # Missing value summary (top 5 with percentages)
        missing_info = []
        missing_counts = df.isna().sum()
        total_rows = max(len(df), 1)
        for col, count in missing_counts.sort_values(ascending=False).head(5).items():
            if count > 0:
                pct = (count / total_rows) * 100
                missing_info.append({'column': col, 'missing_percent': round(pct, 2)})
        metrics['missing_values'] = missing_info

        return metrics

    @staticmethod
    def _count_unique(df: pd.DataFrame, candidate_columns) -> Optional[int]:
        for col in candidate_columns:
            if col in df.columns:
                return int(df[col].nunique(dropna=True))
        return None

    @staticmethod
    def _value_counts(df: pd.DataFrame, candidate_columns) -> Optional[Dict[str, int]]:
        for col in candidate_columns:
            if col in df.columns:
                counts = df[col].value_counts(dropna=True).head(6)
                return {str(idx): int(val) for idx, val in counts.items()}
        return None

    @staticmethod
    def _infer_time_span(df: pd.DataFrame) -> Optional[Dict[str, str]]:
        for col in df.columns:
            if re.search(r'(period|date|month|year)', col, re.IGNORECASE):
                parsed = MetadataCache._parse_date_column(df[col])
                if parsed is not None and parsed.notna().any():
                    valid = parsed.dropna()
                    if not valid.empty:
                        return {
                            'column': col,
                            'start': valid.min().date().isoformat(),
                            'end': valid.max().date().isoformat(),
                            'count': int(valid.nunique())
                        }
        return None

    @staticmethod
    def _parse_date_column(series: pd.Series) -> Optional[pd.Series]:
        s = series.dropna()
        if s.empty:
            return None

        try:
            parsed = pd.to_datetime(s, errors='coerce', infer_datetime_format=True)
            if parsed.notna().any():
                return parsed
        except Exception:
            pass

        # Try numeric codes like 202401 or 20240124
        try:
            as_str = s.astype(str).str.strip().str.replace(r"[^0-9]", "", regex=True)
            if as_str.str.len().between(6, 6).all():
                parsed = pd.to_datetime(as_str, format='%Y%m', errors='coerce')
                if parsed.notna().any():
                    return parsed
            if as_str.str.len().between(8, 8).all():
                parsed = pd.to_datetime(as_str, format='%Y%m%d', errors='coerce')
                if parsed.notna().any():
                    return parsed
        except Exception:
            pass

        return None

    @staticmethod
    def _compute_test_volume(df: pd.DataFrame) -> Optional[Dict[str, Any]]:
        # Identify columns that represent test counts
        test_keywords = ['tested by rdt', 'tested by microscopy', 'testing by rdt', 'testing by microscopy']
        positive_keywords = ['tested positive', 'testing positive']

        test_cols = [col for col in df.columns if any(k in col.lower() for k in test_keywords)]
        positive_cols = [col for col in df.columns if any(k in col.lower() for k in positive_keywords)]

        if not test_cols and not positive_cols:
            return None

        summary: Dict[str, Any] = {}

        if test_cols:
            summary['total_tests'] = int(df[test_cols].select_dtypes(include='number').sum(axis=1).sum())
        if positive_cols:
            summary['total_positive'] = int(df[positive_cols].select_dtypes(include='number').sum(axis=1).sum())

        facility_levels = MetadataCache._value_counts(df, ['FacilityLevel', 'Facility_Level', 'facilitylevel'])
        if facility_levels and test_cols:
            try:
                level_col = next(col for col in ['FacilityLevel', 'Facility_Level', 'facilitylevel'] if col in df.columns)
                grouped = df.groupby(level_col)[test_cols].sum(numeric_only=True).sum(axis=1)
                summary['tests_by_facility_level'] = {str(idx): int(val) for idx, val in grouped.sort_values(ascending=False).items()}
            except Exception:
                pass

        return summary or None

    @staticmethod
    def _build_overview_lines(metadata: Dict[str, Any], metrics: Dict[str, Any]) -> Tuple[str, ...]:
        lines = []

        # Basic data shape (no filename)
        rows = metrics.get('rows')
        rows_str = "Unknown"
        if isinstance(rows, int):
            rows_str = f"~{rows:,}" if metrics.get('rows_estimated') else f"{rows:,}"
        elif rows is not None:
            rows_str = str(rows)

        cols = metrics.get('columns', 'Unknown')
        size_mb = metadata.get('file_size_mb')
        size_info = f" · {size_mb:.1f} MB" if isinstance(size_mb, (int, float)) else ""

        lines.append(f"**Data Shape**: {rows_str} rows × {cols} columns{size_info}")
        lines.append("")  # Blank line

        # Data types section
        dtype_summary = metrics.get('dtype_summary')
        if dtype_summary:
            dtype_parts = []
            for dtype, count in dtype_summary.items():
                if 'int' in dtype or 'float' in dtype:
                    dtype_parts.append(f"{count} numeric")
                elif 'object' in dtype:
                    dtype_parts.append(f"{count} text")
                elif 'datetime' in dtype:
                    dtype_parts.append(f"{count} datetime")
            if dtype_parts:
                lines.append(f"**Data Types**: {', '.join(dtype_parts)}")
                lines.append("")  # Blank line

        # Column information (condensed)
        column_examples = metrics.get('column_examples') or []
        if column_examples and len(column_examples) <= 5:
            lines.append(f"**Columns** ({cols} total): {', '.join(column_examples[:5])}")
        elif column_examples:
            lines.append(f"**Columns** ({cols} total): {', '.join(column_examples[:5])}...")
        lines.append("")  # Blank line

        # Missing data (only if significant)
        if metrics.get('missing_values'):
            significant_missing = [m for m in metrics['missing_values'] if m['missing_percent'] >= 50]
            if significant_missing:
                lines.append("**Note**: Some columns have significant missing data")

        return tuple(lines)
